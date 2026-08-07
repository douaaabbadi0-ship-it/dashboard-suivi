# -*- coding: utf-8 -*-
"""
pdc_parser.py
Parseur du Plan de Charge (PDC) exporte depuis SAP (fichier "EXPORT_*.xlsx").

Structure observee du fichier (feuille unique "Data", ligne 1 = en-tetes) :
Ordre | Avis | Poste technique | Designation | Type d'ordre | Priorite |
Heure Debut calculee | Heure fin calculee | Date de debut planifiee |
Etat immo. | Type de travail | Date planifiee | Statut systeme |
Statut utilisateur | Cree le | Saisi par | Poste travail princ. |
Date de fin planifiee | Centre de couts | Plan d'entretien | Localisation |
Modifie par
"""

from openpyxl import load_workbook
from datetime import datetime, date, time

# TODO (a confirmer avec l'ingenieur) : mapping exact des codes "Type d'ordre".
# Hypothese actuelle basee sur les codes rencontres dans le fichier reel :
#   ZPRV -> Preventif   (le plus frequent, ~1826/2629 lignes)
#   ZCOR -> Correctif   (~365/2629 lignes)
#   ZEST -> inconnu, classe "Autre" par defaut (~399/2629 lignes)
#   ZCRG -> inconnu, classe "Autre" par defaut (~37/2629 lignes)
#   ZARR -> inconnu, classe "Autre" par defaut (~2/2629 lignes)
TYPE_ORDRE_MAPPING = {
    "ZPRV": "Preventif",
    "ZCOR": "Correctif",
    "ZEST": "Autre",
    "ZCRG": "Autre",
    "ZARR": "Autre",
}

HEADER_ROW = 1
DATA_START_ROW = 2

# Index des colonnes (1-based), d'apres l'ordre observe dans le fichier reel.
COLS = {
    "ordre": 1,
    "avis": 2,
    "poste_technique": 3,
    "designation": 4,
    "type_ordre_code": 5,
    "priorite": 6,
    "heure_debut_calc": 7,
    "heure_fin_calc": 8,
    "date_debut_planifiee": 9,
    "etat_immo": 10,
    "type_travail": 11,
    "date_planifiee": 12,
    "statut_systeme": 13,
    "statut_utilisateur": 14,
    "cree_le": 15,
    "saisi_par": 16,
    "poste_travail_princ": 17,
    "date_fin_planifiee": 18,
    "centre_couts": 19,
    "plan_entretien": 20,
    "localisation": 21,
    "modifie_par": 22,
}


def _to_date(value):
    """Normalise une valeur Excel (datetime, date, serial number) en date pure."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        from datetime import timedelta
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    return None


def _clean_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_pdc(filepath, sheet_name="Data"):
    """
    Parse le fichier PDC SAP et retourne une liste de dicts, un par OT.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    rows = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        ordre = ws.cell(row=row, column=COLS["ordre"]).value
        if ordre is None:
            continue

        type_code = _clean_str(ws.cell(row=row, column=COLS["type_ordre_code"]).value)

        rows.append({
            "ordre": str(ordre).strip(),
            "avis": _clean_str(ws.cell(row=row, column=COLS["avis"]).value),
            "poste_technique": _clean_str(ws.cell(row=row, column=COLS["poste_technique"]).value),
            "designation": _clean_str(ws.cell(row=row, column=COLS["designation"]).value),
            "type_ordre_code": type_code,
            "categorie": TYPE_ORDRE_MAPPING.get(type_code, "Autre"),
            "priorite": _clean_str(ws.cell(row=row, column=COLS["priorite"]).value),
            "date_debut_planifiee": _to_date(ws.cell(row=row, column=COLS["date_debut_planifiee"]).value),
            "date_fin_planifiee": _to_date(ws.cell(row=row, column=COLS["date_fin_planifiee"]).value),
            "statut_systeme": _clean_str(ws.cell(row=row, column=COLS["statut_systeme"]).value),
            "statut_utilisateur": _clean_str(ws.cell(row=row, column=COLS["statut_utilisateur"]).value),
            "poste_travail_princ": _clean_str(ws.cell(row=row, column=COLS["poste_travail_princ"]).value),
            "centre_couts": _clean_str(ws.cell(row=row, column=COLS["centre_couts"]).value),
            "localisation": _clean_str(ws.cell(row=row, column=COLS["localisation"]).value),
        })

    return rows


def filter_semaine(pdc_rows, date_debut, date_fin):
    """
    Filtre les OTs du PDC dont la fenetre de planification recoupe
    la periode [date_debut, date_fin] (bornes incluses).

    Un OT est considere "planifie sur la semaine" si son intervalle
    [date_debut_planifiee, date_fin_planifiee] chevauche l'intervalle
    de la semaine. Si une des deux dates de l'OT est manquante, on
    utilise l'autre comme date unique de reference.
    """
    result = []
    for r in pdc_rows:
        d_deb = r["date_debut_planifiee"]
        d_fin = r["date_fin_planifiee"] or d_deb
        d_deb = d_deb or d_fin

        if d_deb is None and d_fin is None:
            continue

        # Chevauchement d'intervalles : [d_deb, d_fin] x [date_debut, date_fin]
        if d_deb <= date_fin and d_fin >= date_debut:
            result.append(r)

    return result


def index_by_ordre(pdc_rows):
    """
    Indexe les lignes du PDC par numero d'OT (cle de rapprochement avec
    le champ 'numero' des rapports journaliers). Un meme Ordre ne devrait
    apparaitre qu'une fois, mais on regroupe en liste par securite.
    """
    index = {}
    for r in pdc_rows:
        index.setdefault(r["ordre"], []).append(r)
    return index