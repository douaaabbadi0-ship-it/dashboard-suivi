# -*- coding: utf-8 -*-
"""
parser.py
Parseur des rapports journaliers JESA (format "Daily Activity Report and Forecast").
"""

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime, date
import re

TEMPS_JOURNALIER_MAROC = 8.8  # heures/jour, temps de travail normal marocain

CRAFT_COLUMNS = {
    "Mecanicien": ("R", "S"),
    "Chaudronnier": ("T", "U"),
    "Soudeur": ("V", "W"),
    "Electricien": ("X", "Y"),
    "Instrumentiste": ("Z", "AA"),
    "Autre": ("AB", "AC"),
}

SECTION_MARKERS = {
    "planned": "PLANNED WORK ORDERS",
    "non_planned": "NON PLANNED WORKS",
    "next_day": "WORKS PLANNED FOR NEXT DAY AND FORECASTS",
}


def _col(letter):
    return column_index_from_string(letter)


def _cell(ws, coord):
    return ws[coord].value


def _find_marker_row(ws, marker_text, max_row=120):
    for row in range(1, max_row):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() == marker_text:
            return row
    return None


def _parse_date(raw_value):
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    if isinstance(raw_value, (int, float)):
        from datetime import timedelta
        return (datetime(1899, 12, 30) + timedelta(days=float(raw_value))).date()
    if isinstance(raw_value, str):
        raw_value = raw_value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw_value, fmt).date()
            except ValueError:
                continue
    return None


def _extract_craft_hours(ws, row):
    result = {}
    for craft, (est_letter, real_letter) in CRAFT_COLUMNS.items():
        est_val = _cell(ws, f"{est_letter}{row}")
        real_val = _cell(ws, f"{real_letter}{row}")
        if est_val is not None or real_val is not None:
            try:
                est_val = float(est_val) if est_val is not None else 0.0
            except (TypeError, ValueError):
                est_val = 0.0
            try:
                real_val = float(real_val) if real_val is not None else 0.0
            except (TypeError, ValueError):
                real_val = 0.0
            if est_val or real_val:
                result[craft] = {"estime": est_val, "reel": real_val}
    return result


def _parse_table_section(ws, start_marker_row, end_marker_row=None, max_row_limit=80):
    """
    Parse une section de tableau (PLANNED / NON PLANNED / NEXT DAY).

    IMPORTANT (fix bug) : la lecture est desormais bornee explicitement par
    `end_marker_row`, c'est-a-dire la ligne ou commence la section SUIVANTE
    (quand elle est connue). Avant ce fix, la fonction se fiait uniquement a
    la detection de 2 lignes vides consecutives pour savoir ou s'arreter.
    Si l'espacement reel entre deux tableaux dans le fichier Excel ne
    contenait pas exactement 2 lignes vides d'affilee (une seule ligne vide,
    une ligne de sous-titre, etc.), le scan continuait au-dela de la fin
    reelle du tableau et venait "avaler" la ligne d'en-tete de la section
    suivante (ex: la ligne "NON PLANNED WORKS") comme si c'etait un OT
    planifie valide, avec un statut vide. Consequence concrete observee :
    nb_planifie gonfle artificiellement (ex: 141 au lieu de 130), ce qui fait
    chuter le taux d'execution calcule (nb_executes / nb_planifie) alors que
    la realite terrain est un taux de 100%.

    On ne peut donc plus jamais depasser `end_marker_row - 1`, quel que soit
    le nombre de lignes vides rencontrees.
    """
    if start_marker_row is None:
        return []

    data_start = start_marker_row + 3

    if end_marker_row is not None:
        hard_limit = end_marker_row - 1
    else:
        hard_limit = start_marker_row + max_row_limit

    rows = []
    row = data_start
    empty_streak = 0

    while row < hard_limit:
        num = _cell(ws, f"A{row}")
        type_ = _cell(ws, f"B{row}")
        tag = _cell(ws, f"C{row}")
        desc = _cell(ws, f"E{row}")
        exec_flag = _cell(ws, f"L{row}")
        comment = _cell(ws, f"N{row}")

        if num is None and type_ is None and tag is None and desc is None:
            empty_streak += 1
            if empty_streak >= 2:
                break
        else:
            empty_streak = 0
            crafts = _extract_craft_hours(ws, row)
            rows.append({
                "numero": num,
                "type": (str(type_).strip() if type_ else None),
                "tag_equipement": (str(tag).strip() if tag else None),
                "description": (str(desc).strip() if desc else None),
                "statut": (str(exec_flag).strip() if exec_flag else None),
                "commentaire": (str(comment).strip() if comment else None),
                "heures": crafts,
            })
        row += 1

    return rows


def parse_daily_report(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    date_val = None
    for col_letter in ("V", "W", "X", "Y", "Z"):
        candidate = _parse_date(_cell(ws, f"{col_letter}6"))
        if candidate:
            date_val = candidate
            break

    meta = {
        "projet": _cell(ws, "G5"),
        "client": _cell(ws, "V5"),
        "numero_projet": _cell(ws, "G6"),
        "date": date_val,
        "responsable": (str(_cell(ws, "G7")).strip() if _cell(ws, "G7") else None),
        "unite": _cell(ws, "V7"),
        "fichier_source": filepath,
    }

    planned_row = _find_marker_row(ws, SECTION_MARKERS["planned"])
    non_planned_row = _find_marker_row(ws, SECTION_MARKERS["non_planned"])
    next_day_row = _find_marker_row(ws, SECTION_MARKERS["next_day"])

    # Chaque section est desormais bornee par le debut de la section
    # suivante (fix bug de debordement de scan, voir _parse_table_section).
    planifie = _parse_table_section(ws, planned_row, end_marker_row=non_planned_row)
    non_planifie = _parse_table_section(ws, non_planned_row, end_marker_row=next_day_row)
    prevu_lendemain = _parse_table_section(ws, next_day_row, end_marker_row=None)

    totaux_heures = {}
    for section in (planifie, non_planifie):
        for item in section:
            for craft, vals in item["heures"].items():
                if craft not in totaux_heures:
                    totaux_heures[craft] = {"estime": 0.0, "reel": 0.0}
                totaux_heures[craft]["estime"] += vals["estime"]
                totaux_heures[craft]["reel"] += vals["reel"]

    nb_planifie = len(planifie)
    nb_executes = sum(1 for i in planifie if i["statut"] and i["statut"].lower().startswith("y"))
    nb_preventif = sum(1 for i in planifie + non_planifie if i["type"] and "prevent" in i["type"].lower())
    nb_correctif = sum(1 for i in planifie + non_planifie if i["type"] and "correct" in i["type"].lower())
    total_heures_reel = sum(v["reel"] for v in totaux_heures.values())
    total_heures_estime = sum(v["estime"] for v in totaux_heures.values())

    kpis = {
        "nb_planifie": nb_planifie,
        "nb_executes": nb_executes,
        "taux_execution": round(100 * nb_executes / nb_planifie, 1) if nb_planifie else None,
        "nb_non_planifie": len(non_planifie),
        "nb_preventif": nb_preventif,
        "nb_correctif": nb_correctif,
        "total_heures_reel": round(total_heures_reel, 2),
        "total_heures_estime": round(total_heures_estime, 2),
        "efficacite_temps": round(100 * total_heures_reel / total_heures_estime, 1) if total_heures_estime else None,
    }

    return {
        "meta": meta,
        "planifie": planifie,
        "non_planifie": non_planifie,
        "prevu_lendemain": prevu_lendemain,
        "totaux_heures": totaux_heures,
        "kpis": kpis,
    }


def aggregate_week(daily_reports):
    daily_reports = sorted(daily_reports, key=lambda r: r["meta"]["date"] or date.min)

    heures_semaine = {}
    total_planifie = 0
    total_executes = 0
    total_non_planifie = 0
    total_preventif = 0
    total_correctif = 0
    total_heures_reel = 0.0
    total_heures_estime = 0.0

    interventions_correctives = []
    interventions_urgentes = []
    toutes_interventions_preventives = []

    for r in daily_reports:
        d = r["meta"]["date"]
        for craft, vals in r["totaux_heures"].items():
            if craft not in heures_semaine:
                heures_semaine[craft] = {"estime": 0.0, "reel": 0.0}
            heures_semaine[craft]["estime"] += vals["estime"]
            heures_semaine[craft]["reel"] += vals["reel"]

        total_planifie += r["kpis"]["nb_planifie"]
        total_executes += r["kpis"]["nb_executes"] or 0
        total_non_planifie += r["kpis"]["nb_non_planifie"]
        total_preventif += r["kpis"]["nb_preventif"]
        total_correctif += r["kpis"]["nb_correctif"]
        total_heures_reel += r["kpis"]["total_heures_reel"]
        total_heures_estime += r["kpis"]["total_heures_estime"]

        for item in r["planifie"]:
            if item["type"] and "prevent" in item["type"].lower():
                toutes_interventions_preventives.append({**item, "date": d})

        for item in r["non_planifie"]:
            interventions_correctives.append({**item, "date": d})
            if item["statut"] and item["statut"].lower().startswith("y"):
                interventions_urgentes.append({**item, "date": d})

    taux_execution_semaine = round(100 * total_executes / total_planifie, 1) if total_planifie else None
    efficacite_temps = round(100 * total_heures_reel / total_heures_estime, 1) if total_heures_estime else None

    dates = [r["meta"]["date"] for r in daily_reports if r["meta"]["date"]]
    # nb_jours = nombre de JOURS CALENDAIRES UNIQUES, pas nombre de fichiers
    # uploades. Avec un seul rapport par jour (cas mono-metier actuel), les
    # deux coincident. Mais si un jour plusieurs superviseurs de metiers
    # differents (elec + meca, etc.) uploadent chacun leur rapport pour les
    # memes dates, len(daily_reports) doublerait/triplerait artificiellement
    # nb_jours (ex: 10 fichiers pour 5 jours calendaires reels si 2 metiers
    # sont fusionnes). Ce nb_jours gonfle est ensuite utilise pour calculer
    # la capacite theorique dans _compute_occupancy_rate (effectif x 8.8h x
    # nb_jours), ce qui sous-evaluerait artificiellement l'Occupancy Rate.
    # On utilise donc le nombre de dates distinctes, robuste peu importe le
    # nombre de fichiers/metiers fusionnes sur la meme periode.
    nb_jours_uniques = len(set(dates)) if dates else 0

    return {
        "periode": {
            "debut": min(dates) if dates else None,
            "fin": max(dates) if dates else None,
            "nb_jours": nb_jours_uniques,
        },
        "projet": daily_reports[0]["meta"] if daily_reports else {},
        "heures_semaine": heures_semaine,
        "kpis": {
            "total_planifie": total_planifie,
            "total_executes": total_executes,
            "taux_execution": taux_execution_semaine,
            "total_non_planifie": total_non_planifie,
            "total_preventif": total_preventif,
            "total_correctif": total_correctif,
            "total_heures_reel": round(total_heures_reel, 2),
            "total_heures_estime": round(total_heures_estime, 2),
            "efficacite_temps": efficacite_temps,
        },
        "interventions_correctives": interventions_correctives,
        "interventions_urgentes": interventions_urgentes,
        "interventions_preventives": toutes_interventions_preventives,
        "rapports_journaliers": daily_reports,
    }


def compute_kpis_officiels(week_data, effectifs_metier=None):
    """
    Calcule les KPIs officiels de la procedure JESA-MS-AM-MX-PRO-0100 (Appendix A),
    dans la limite de ce que permettent les rapports journaliers actuels.

    effectifs_metier : dict optionnel {metier: effectif} pour la ligne concernee
                        (voir effectifs_kpi.get_effectifs_ligne). Si fourni, permet
                        de calculer l'Occupancy Rate.
    """
    k = week_data["kpis"]
    heures = week_data["heures_semaine"]

    wrench_time_par_metier = {}
    for craft, v in heures.items():
        wrench_time_par_metier[craft] = round(100 * v["reel"] / v["estime"], 1) if v["estime"] else None

    # MTTR approché : heures réelles des interventions CORRECTIVES uniquement / nb correctifs
    heures_correctif_reel = sum(
        sum(vals["reel"] for vals in item["heures"].values())
        for item in week_data["interventions_correctives"]
    )
    mttr_approx = (
        round(heures_correctif_reel / k["total_non_planifie"], 2)
        if k["total_non_planifie"] else None
    )

    occupancy_rate = _compute_occupancy_rate(week_data, effectifs_metier)

    return {
        "wrench_time_par_metier": wrench_time_par_metier,
        "schedule_compliance": k["taux_execution"],
        "mh_compliance_global": k["efficacite_temps"],
        "mttr_approx_heures": mttr_approx,
        "rate_of_rework": None,
        "wo_backlog": None,
        "wo_non_compliance": None,
        "occupancy_rate": occupancy_rate,
    }


def _compute_occupancy_rate(week_data, effectifs_metier):
    """
    Occupancy Rate = heures reelles travaillees / capacite theorique disponible.
    Capacite theorique (par metier) = effectif x 8.8h/jour x nb_jours de la periode.

    Retourne None si aucun effectif n'est fourni pour la ligne (cas non
    encore configure via la page /effectifs).
    """
    if not effectifs_metier:
        return None

    nb_jours = week_data["periode"]["nb_jours"] or 0
    if not nb_jours:
        return None

    heures = week_data["heures_semaine"]

    occupancy_par_metier = {}
    total_reel = 0.0
    total_capacite = 0.0

    for craft, vals in heures.items():
        effectif = effectifs_metier.get(craft, 0) or 0
        if effectif <= 0:
            occupancy_par_metier[craft] = None
            continue

        capacite = effectif * TEMPS_JOURNALIER_MAROC * nb_jours
        occupancy_par_metier[craft] = round(100 * vals["reel"] / capacite, 1) if capacite else None
        total_reel += vals["reel"]
        total_capacite += capacite

    occupancy_global = round(100 * total_reel / total_capacite, 1) if total_capacite else None

    return {
        "global": occupancy_global,
        "par_metier": occupancy_par_metier,
    }